#!/usr/bin/python


import sys, os, warnings

warnings.simplefilter("ignore")
from openpyxl import load_workbook      #To read excel file imported from openpyxl library
reference = {}  #Initialize Dictionary
iupac = {"AA":"A", "GG":"G", "CC":"C", "TT":"T", "AG":"R", "TC":"Y", "AT":"W", "GC":"S", "TG":"K", "AC":"M", "--":"N"}  #Dictionary for IUPAC Naming Convention

os.rename(sys.argv[1],sys.argv[1] + ".xlsx")    #rename the .dat file from galaxy database to have an extension of .xlsx to let the program read it.
excel_file = load_workbook(sys.argv[1] + ".xlsx")       #read the excel file
multi_fasta = open(sys.argv[2], 'w')    #opens the output file for writing
reference_file = open(sys.argv[3])      #opens the reference file for reference

for line in reference_file:
        reference[str(line.split()[2])] = True  #assigns possible positions to the reference dictionary

sheet = excel_file[excel_file.get_sheet_names()[0]]     #reads the excel file

row_count = sheet.max_row       #gets the row count
col_count = sheet.max_column    #gets the column count



for i in range(7,row_count + 1):
        multi_fasta.write(">" + str(sheet.cell(row=i, column=4).value).replace(" ","_") + "_" + str(sheet.cell(row=i, column=6).value) + "\n")   #writes the header of the fasta
        for j in range(14,col_count + 1):
                position = str(sheet.cell(row=3, column=j).value)
                if reference.get(position):     #if position is true
                        multi_fasta.write(iupac[str(sheet.cell(row=i, column=j).value)])        #decode and write it using the iupac naming convention
        multi_fasta.write("\n") #writes a newline 
        
        
os.rename(sys.argv[1] + ".xlsx",sys.argv[1])    #renames the excel file back to .dat file
